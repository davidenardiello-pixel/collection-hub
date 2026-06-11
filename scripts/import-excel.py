#!/usr/bin/env python3
"""Import bookings and expenses from File rendita speciale FY26.xlsx."""

from __future__ import annotations

import calendar
import json
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / "File rendita speciale FY26.xlsx"
OUTPUT_PATH = ROOT / "src/lib/excel-data.json"
FISCAL_YEAR = 2026
INCOME_SHEET = "INCOME BACK-END"
EXPENSE_SHEET = "EXPENSES BACK-END"
INCOME_SCAN_ROWS = 9400
EXPENSE_SCAN_ROWS = 9400

PROPERTY_MAP = {
    "Regina Cappellari": "regina-cappellari",
    "Mistangelo": "mistangelo",
    "Re Di Roma": "re-di-roma",
    "Cipro": "cipro",
    "Palombini": "palombini",
    "Porto Fluviale": "porto-fluviale",
    "Gregorio VII": "gregorio-vii",
    "Hallo Claudia": "hallo-claudia",
}

PLATFORM_MAP = {
    "Booking": "booking",
    "Airbnb": "airbnb",
    "Diretta": "diretta",
    "VRBO": "vrbo",
    "Prenotazione diretta": "prenotazione-diretta",
    "Rome Collection": "rome-collection",
}

CATEGORY_MAP = {
    "Utenze": "utenze",
    "KrossBooking": "krossbooking",
    "Pulizie": "pulizie",
    "Wi-fi": "wifi",
    "Tassa rifiuti": "tassa-rifiuti",
    "Tassa acqua": "tassa-acqua",
    "Biancheria": "biancheria",
    "Arredamento": "arredamento",
    "Condominio": "condominio",
    "Mutuo": "mutuo",
    "Gas": "gas",
    "Revenue Managment": "revenue-management",
    "Assicurazioni": "assicurazioni",
    "manutenzione": "manutenzione",
    "Affitto": "affitto",
    "Com. Booking": "com-booking",
    "Com.Airbnb": "com-airbnb",
    "IVA": "iva",
    "Lavanderia": "lavanderia",
}

EXPECTED_TOTALS = {
    1: (18099.61, 16317.85),
    2: (16153.05, 15077.98),
    3: (25130.55, 16168.30),
    4: (30673.38, 17457.42),
    5: (39688.00, 22541.66),
}

EXPECTED_PROPERTY = {
    "cipro": {
        1: (3914.90, 2570.13),
        2: (2791.09, 2176.00),
        3: (3829.81, 2346.00),
        4: (4323.24, 2402.10),
        5: (3462.74, 2093.14),
    },
    "gregorio-vii": {
        5: (6970.82, 5035.02),
    },
    "hallo-claudia": {
        1: (950.00, 390.00),
        2: (820.00, 250.00),
    },
    "mistangelo": {
        1: (2602.79, 2379.82),
        2: (2501.40, 2177.84),
        3: (5015.58, 3386.33),
        4: (5863.70, 3601.67),
        5: (8770.59, 4313.41),
    },
    "palombini": {
        1: (1553.00, 2699.72),
        2: (2123.20, 2332.49),
        3: (2983.97, 1916.00),
        4: (3462.13, 2326.42),
        5: (2150.00, 1600.00),
    },
    "porto-fluviale": {
        1: (2069.22, 2524.00),
        2: (1879.77, 2211.00),
        3: (5257.62, 2959.43),
        4: (5920.90, 3025.01),
        5: (7264.68, 2970.70),
    },
    "re-di-roma": {
        1: (3220.39, 2264.78),
        2: (1862.05, 2091.10),
        3: (2653.14, 2098.51),
        4: (3835.37, 2244.54),
        5: (3117.47, 2041.06),
    },
    "regina-cappellari": {
        1: (3789.31, 3489.40),
        2: (4175.54, 3839.55),
        3: (5390.43, 3462.03),
        4: (7268.04, 3857.68),
        5: (7951.70, 4488.33),
    },
}


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    ascii_text = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
    slug = re.sub(r"[^a-z0-9]+", "-", ascii_text.lower()).strip("-")
    return slug or "item"


def to_iso_date(value: date | datetime | str | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text[:10] if text else None


def to_float(value: object) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


def month_bounds(year: int, month: int) -> tuple[str, str]:
    last_day = calendar.monthrange(year, month)[1]
    return (
        f"{year:04d}-{month:02d}-01",
        f"{year:04d}-{month:02d}-{last_day:02d}",
    )


def map_property(name: str | None) -> str:
    if not name:
        raise ValueError("Missing property name")
    return PROPERTY_MAP.get(str(name).strip(), slugify(str(name)))


def map_platform(name: str | None) -> str:
    if not name:
        raise ValueError("Missing platform name")
    return PLATFORM_MAP.get(str(name).strip(), slugify(str(name)))


def map_category(name: str | None) -> str:
    if not name:
        raise ValueError("Missing category name")
    return CATEGORY_MAP.get(str(name).strip(), slugify(str(name)))


def import_bookings(workbook: openpyxl.Workbook) -> list[dict]:
    worksheet = workbook[INCOME_SHEET]
    bookings: list[dict] = []

    for row in range(7, INCOME_SCAN_ROWS + 1):
        property_name = worksheet.cell(row, 3).value
        if not property_name:
            continue

        total = to_float(worksheet.cell(row, 10).value)
        if total == 0:
            continue

        year = worksheet.cell(row, 13).value
        month = worksheet.cell(row, 12).value
        if year != FISCAL_YEAR or not month:
            continue

        month = int(month)
        check_in = to_iso_date(worksheet.cell(row, 6).value)
        check_out = to_iso_date(worksheet.cell(row, 7).value)
        if not check_in or not check_out:
            check_in, check_out = month_bounds(FISCAL_YEAR, month)

        bookings.append(
            {
                "description": str(worksheet.cell(row, 2).value or "Totale").strip(),
                "propertyId": map_property(property_name),
                "platformId": map_platform(worksheet.cell(row, 4).value),
                "checkIn": check_in,
                "checkOut": check_out,
                "grossIncome": round(to_float(worksheet.cell(row, 8).value), 2),
                "cleaningFee": round(to_float(worksheet.cell(row, 9).value), 2),
                "otaCommission": 0,
                "legacyIncomeAttribution": True,
                "importedFromExcel": True,
            }
        )

    return bookings


def import_expenses(workbook: openpyxl.Workbook) -> list[dict]:
    worksheet = workbook[EXPENSE_SHEET]
    expenses: list[dict] = []

    for row in range(7, EXPENSE_SCAN_ROWS + 1):
        property_name = worksheet.cell(row, 3).value
        if not property_name:
            continue

        amount = to_float(worksheet.cell(row, 6).value)
        if amount == 0:
            continue

        year = worksheet.cell(row, 9).value
        month = worksheet.cell(row, 8).value
        if year != FISCAL_YEAR or not month:
            continue

        expense_date = to_iso_date(worksheet.cell(row, 2).value)
        if not expense_date:
            _, expense_date = month_bounds(FISCAL_YEAR, int(month))

        notes = worksheet.cell(row, 7).value
        expenses.append(
            {
                "date": expense_date,
                "propertyId": map_property(property_name),
                "categoryId": map_category(worksheet.cell(row, 4).value),
                "description": str(worksheet.cell(row, 5).value or "").strip(),
                "amount": round(amount, 2),
                "importedFromExcel": True,
                **({"notes": str(notes).strip()} if notes else {}),
            }
        )

    return expenses


def summarize(bookings: list[dict], expenses: list[dict]) -> dict[int, dict[str, float]]:
    summary: dict[int, dict[str, float]] = {}

    for month in range(1, 13):
        income = sum(
            booking["grossIncome"] + booking["cleaningFee"]
            for booking in bookings
            if int(booking["checkIn"][5:7]) == month
        )
        expense_total = sum(
            expense["amount"]
            for expense in expenses
            if int(expense["date"][5:7]) == month
        )
        if income or expense_total:
            summary[month] = {
                "income": round(income, 2),
                "expenses": round(expense_total, 2),
                "profit": round(income - expense_total, 2),
            }

    return summary


def summarize_property(
    bookings: list[dict],
    expenses: list[dict],
    property_id: str,
) -> dict[int, dict[str, float]]:
    summary: dict[int, dict[str, float]] = {}

    for month in range(1, 13):
        income = sum(
            booking["grossIncome"] + booking["cleaningFee"]
            for booking in bookings
            if booking["propertyId"] == property_id
            and int(booking["checkIn"][5:7]) == month
        )
        expense_total = sum(
            expense["amount"]
            for expense in expenses
            if expense["propertyId"] == property_id
            and int(expense["date"][5:7]) == month
        )
        if income or expense_total:
            summary[month] = {
                "income": round(income, 2),
                "expenses": round(expense_total, 2),
                "profit": round(income - expense_total, 2),
            }

    return summary


def verify_import(bookings: list[dict], expenses: list[dict]) -> None:
    monthly = summarize(bookings, expenses)

    for month, (expected_income, expected_expenses) in EXPECTED_TOTALS.items():
        actual = monthly.get(month)
        if not actual:
            raise SystemExit(f"Missing totals for month {month}")

        if abs(actual["income"] - expected_income) > 0.05:
            raise SystemExit(
                f"Month {month} income mismatch: {actual['income']} != {expected_income}"
            )
        if abs(actual["expenses"] - expected_expenses) > 0.05:
            raise SystemExit(
                f"Month {month} expenses mismatch: {actual['expenses']} != {expected_expenses}"
            )

    for property_id, months in EXPECTED_PROPERTY.items():
        property_monthly = summarize_property(bookings, expenses, property_id)
        for month, (expected_income, expected_expenses) in months.items():
            actual = property_monthly.get(month)
            if not actual:
                raise SystemExit(
                    f"Missing {property_id} data for month {month}"
                )
            if abs(actual["income"] - expected_income) > 0.05:
                raise SystemExit(
                    f"{property_id} month {month} income mismatch: "
                    f"{actual['income']} != {expected_income}"
                )
            if abs(actual["expenses"] - expected_expenses) > 0.05:
                raise SystemExit(
                    f"{property_id} month {month} expenses mismatch: "
                    f"{actual['expenses']} != {expected_expenses}"
                )


def main() -> None:
    if not EXCEL_PATH.exists():
        raise SystemExit(f"Excel file not found: {EXCEL_PATH}")

    workbook = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    bookings = import_bookings(workbook)
    expenses = import_expenses(workbook)
    verify_import(bookings, expenses)

    payload = {"bookings": bookings, "expenses": expenses}
    OUTPUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    monthly = summarize(bookings, expenses)
    total_income = round(sum(month["income"] for month in monthly.values()), 2)
    total_expenses = round(sum(month["expenses"] for month in monthly.values()), 2)

    print(f"Wrote {len(bookings)} bookings and {len(expenses)} expenses to {OUTPUT_PATH}")
    print(f"YTD income: {total_income:.2f} | expenses: {total_expenses:.2f} | profit: {total_income - total_expenses:.2f}")
    for month, values in sorted(monthly.items()):
        print(
            f"  M{month:02d}: income {values['income']:.2f} | "
            f"expenses {values['expenses']:.2f} | profit {values['profit']:.2f}"
        )

    regina = summarize_property(bookings, expenses, "regina-cappellari")
    print("Regina Cappellari:")
    for month, values in sorted(regina.items()):
        print(
            f"  M{month:02d}: income {values['income']:.2f} | "
            f"expenses {values['expenses']:.2f} | profit {values['profit']:.2f}"
        )


if __name__ == "__main__":
    main()
